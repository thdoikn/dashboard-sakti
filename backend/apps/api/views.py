"""
REST API viewsets and function-based views for the SAKTI-OIKN dashboard.

Endpoints:
  GET  /api/satker/                         — list + filter by aktif
  POST/PUT/DELETE /api/satker/{id}/         — satker CRUD
  GET  /api/referensi/                      — read-only reference data
  GET  /api/anggaran/                       — budget data (filterable)
  GET  /api/realisasi/                      — disbursement data (filterable)
  GET  /api/capaian-ro/                     — performance achievement data
  GET  /api/sync-log/                       — audit log of all sync task runs
  GET  /api/sync-status/                    — latest sync result for dashboard badge
  GET  /api/export/excel/                   — download .xlsx file with filters
  GET  /api/users/                          — user list sorted by role + last login (auth required)
"""

import io
from collections import defaultdict
from datetime import timedelta
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import filters, permissions, viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response

from apps.anggaran.models import Anggaran, Referensi
from apps.capaian.models import CapaianRO
from apps.realisasi.models import Realisasi
from apps.satker.models import Satker
from apps.sync_log.models import SyncLog

from .filters import AnggaranFilter, CapaianROFilter, RealisasiFilter, SyncLogFilter
from .serializers import (
    AnggaranSerializer,
    CapaianROSerializer,
    ReferensiSerializer,
    RealisasiSerializer,
    SatkerSerializer,
    SyncLogSerializer,
)


# ---------------------------------------------------------------------------
# Satker viewset (full CRUD)
# ---------------------------------------------------------------------------


class SatkerViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for Satker records.

    Supports filtering by ?aktif=true/false query param.
    No authentication is required in v1 (internal-only via network restriction).
    """

    serializer_class = SatkerSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["aktif", "kode_kppn", "kode_kementerian"]
    search_fields = ["kode_satker", "nama_satker"]
    ordering_fields = ["kode_satker", "nama_satker", "created_at"]
    ordering = ["kode_satker"]

    def get_queryset(self):
        return Satker.objects.all()


# ---------------------------------------------------------------------------
# Read-only viewsets
# ---------------------------------------------------------------------------


class ReferensiViewSet(viewsets.ReadOnlyModelViewSet):
    """Read-only access to the SAKTI reference code lookup table."""

    serializer_class = ReferensiSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ["jenis", "kode"]
    search_fields = ["kode", "uraian"]

    def get_queryset(self):
        return Referensi.objects.all()


class AnggaranViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Budget data endpoint.

    Supports filtering by satker (id or kode_satker) and tahun_anggaran.
    The ?granularity=daily|monthly param is accepted but does not affect
    the row-level data here — aggregation is left to the frontend for now,
    since budget data is already per-item (not time-series).
    """

    serializer_class = AnggaranSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = AnggaranFilter
    ordering_fields = ["tahun_anggaran", "kode_program", "kode_kegiatan", "total"]
    ordering = ["-tahun_anggaran", "kode_program"]

    def get_queryset(self):
        return Anggaran.objects.select_related("satker").all()


class RealisasiViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Disbursement data endpoint.

    Supports date range filtering via tgl_sp2d_after / tgl_sp2d_before,
    plus satker and tahun_anggaran filters.
    """

    serializer_class = RealisasiSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = RealisasiFilter
    ordering_fields = ["tgl_sp2d", "tgl_spm", "nilai_sp2d", "tahun_anggaran"]
    ordering = ["-tgl_sp2d"]

    def get_queryset(self):
        return Realisasi.objects.select_related("satker").all()


class CapaianROViewSet(viewsets.ReadOnlyModelViewSet):
    """Performance achievement data endpoint."""

    serializer_class = CapaianROSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = CapaianROFilter
    ordering_fields = ["kode_periode", "sub_output_kode"]
    ordering = ["-kode_periode"]

    def get_queryset(self):
        return CapaianRO.objects.select_related("satker").all()


class SyncLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Audit log of all Airflow sync task runs."""

    serializer_class = SyncLogSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = SyncLogFilter
    ordering_fields = ["started_at", "finished_at", "status"]
    ordering = ["-started_at"]

    def get_queryset(self):
        return SyncLog.objects.select_related("satker").all()


# ---------------------------------------------------------------------------
# Function-based views
# ---------------------------------------------------------------------------


@api_view(["GET"])
def sync_status_view(request):
    """
    Returns the most recent SyncLog entry for each unique dag_run_id,
    plus a summary of overall sync health for the dashboard status badge.

    Response shape:
      {
        "last_sync": { ...SyncLog fields... } | null,
        "last_successful_sync": { ...SyncLog fields... } | null,
        "timestamp": "ISO-8601"
      }
    """
    last_log = SyncLog.objects.select_related("satker").order_by("-started_at").first()
    last_success = (
        SyncLog.objects.filter(status=SyncLog.STATUS_SUCCESS)
        .select_related("satker")
        .order_by("-started_at")
        .first()
    )

    return Response(
        {
            "last_sync": SyncLogSerializer(last_log).data if last_log else None,
            "last_successful_sync": (
                SyncLogSerializer(last_success).data if last_success else None
            ),
            "timestamp": timezone.now().isoformat(),
        }
    )


@api_view(["GET"])
def sync_history_view(request):
    """
    Returns a list of DAG runs aggregated from SyncLog records, newest first.

    Each entry groups all SyncLog rows that share a dag_run_id and computes:
      overall_status  — success (all ok) / partial (some failed) / failed (all failed)
      success_count, failed_count, total_tasks, total_rows
      tasks           — per-task detail list

    Query params:
      days  — how many days of history to include (default 30, max 90)
    """
    days = min(int(request.query_params.get("days", 30)), 90)
    since = timezone.now() - timedelta(days=days)

    logs = (
        SyncLog.objects.filter(started_at__gte=since)
        .select_related("satker")
        .order_by("dag_run_id", "started_at")
    )

    # Group rows by dag_run_id in Python (cheaper than a GROUP BY for small datasets)
    grouped: dict[str, list] = defaultdict(list)
    for log in logs:
        grouped[log.dag_run_id].append(log)

    result = []
    for dag_run_id, tasks in sorted(
        grouped.items(), key=lambda kv: kv[1][0].started_at, reverse=True
    ):
        started_at   = min(t.started_at for t in tasks)
        finish_times = [t.finished_at for t in tasks if t.finished_at]
        finished_at  = max(finish_times) if finish_times else None

        success_count = sum(1 for t in tasks if t.status == SyncLog.STATUS_SUCCESS)
        failed_count  = sum(1 for t in tasks if t.status == SyncLog.STATUS_FAILED)
        total_rows    = sum(t.row_count or 0 for t in tasks)

        if failed_count == 0:
            overall_status = "success"
        elif success_count == 0:
            overall_status = "failed"
        else:
            overall_status = "partial"

        duration_s = (
            int((finished_at - started_at).total_seconds()) if finished_at else None
        )

        task_list = []
        for t in sorted(tasks, key=lambda x: x.started_at):
            t_dur = (
                int((t.finished_at - t.started_at).total_seconds())
                if t.finished_at else None
            )
            task_list.append({
                "task_name":       t.task_name,
                "status":          t.status,
                "satker_nama":     t.satker.nama_satker if t.satker else None,
                "row_count":       t.row_count,
                "error_message":   t.error_message,
                "started_at":      t.started_at.isoformat(),
                "finished_at":     t.finished_at.isoformat() if t.finished_at else None,
                "duration_seconds": t_dur,
            })

        result.append({
            "dag_run_id":       dag_run_id,
            "date":             started_at.date().isoformat(),
            "started_at":       started_at.isoformat(),
            "finished_at":      finished_at.isoformat() if finished_at else None,
            "duration_seconds": duration_s,
            "overall_status":   overall_status,
            "total_tasks":      len(tasks),
            "success_count":    success_count,
            "failed_count":     failed_count,
            "total_rows":       total_rows,
            "tasks":            task_list,
        })

    return Response(result)


@api_view(["GET"])
def excel_export_view(request):
    """
    Generate and return an Excel (.xlsx) file containing Anggaran and Realisasi data.

    Query params:
      satker          — filter by satker id (integer)
      kode_satker     — filter by satker kode_satker (string)
      tahun_anggaran  — filter by fiscal year (integer)
      granularity     — 'daily' or 'monthly' (informational only, accepted but not
                        applied to row selection — aggregation is handled in the frontend)

    The file has two sheets:
      - "Anggaran"   with columns: Kode Satker, Nama Satker, Kode Program,
                     Kode Kegiatan, Kode Output, Kode Item, Uraian Item, Total (Rp)
      - "Realisasi"  with columns: Kode Satker, Nama Satker, No SPP, No SPM,
                     No SP2D, Tgl SP2D, Nilai SP2D (Rp), Status
    """
    # --- Parse filters ---
    satker_id = request.query_params.get("satker")
    kode_satker = request.query_params.get("kode_satker")
    tahun_anggaran = request.query_params.get("tahun_anggaran")

    anggaran_qs = Anggaran.objects.select_related("satker").all()
    realisasi_qs = Realisasi.objects.select_related("satker").all()

    if satker_id:
        anggaran_qs = anggaran_qs.filter(satker__id=satker_id)
        realisasi_qs = realisasi_qs.filter(satker__id=satker_id)
    if kode_satker:
        anggaran_qs = anggaran_qs.filter(satker__kode_satker=kode_satker)
        realisasi_qs = realisasi_qs.filter(satker__kode_satker=kode_satker)
    if tahun_anggaran:
        anggaran_qs = anggaran_qs.filter(tahun_anggaran=tahun_anggaran)
        realisasi_qs = realisasi_qs.filter(tahun_anggaran=tahun_anggaran)

    # --- Build workbook ---
    wb = Workbook()

    # Style helpers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_format = "#,##0.00"

    def _write_sheet(ws, headers, rows, currency_col_indices):
        """Write headers + rows to a worksheet with consistent formatting."""
        # Header row
        ws.append(headers)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row_data in rows:
            ws.append(row_data)

        # Apply currency format and auto-size columns
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            # Number format for currency columns
            if col_idx in currency_col_indices:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = currency_format
            # Column width: derive from header length or fixed minimum
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, min(ws.max_row + 1, 50)):  # sample first 50 rows
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

    # --- Sheet 1: Anggaran ---
    ws_ang = wb.active
    ws_ang.title = "Anggaran"

    ang_headers = [
        "Kode Satker",
        "Nama Satker",
        "Kode Program",
        "Kode Kegiatan",
        "Kode Output",
        "Kode Item",
        "Uraian Item",
        "Total (Rp)",
    ]
    ang_rows = [
        [
            row.satker.kode_satker,
            row.satker.nama_satker,
            row.kode_program,
            row.kode_kegiatan,
            row.kode_output,
            row.kode_item,
            row.uraian_item,
            float(row.total) if row.total is not None else None,
        ]
        for row in anggaran_qs
    ]
    # Column 8 (Total) is currency
    _write_sheet(ws_ang, ang_headers, ang_rows, currency_col_indices={8})

    # --- Sheet 2: Realisasi ---
    ws_real = wb.create_sheet(title="Realisasi")

    real_headers = [
        "Kode Satker",
        "Nama Satker",
        "No SPP",
        "No SPM",
        "No SP2D",
        "Tgl SP2D",
        "Nilai SP2D (Rp)",
        "Status",
    ]
    real_rows = [
        [
            row.satker.kode_satker,
            row.satker.nama_satker,
            row.no_spp,
            row.no_spm,
            row.no_sp2d,
            row.tgl_sp2d.isoformat() if row.tgl_sp2d else None,
            float(row.nilai_sp2d) if row.nilai_sp2d is not None else None,
            row.status_data,
        ]
        for row in realisasi_qs
    ]
    # Column 7 (Nilai SP2D) is currency
    _write_sheet(ws_real, real_headers, real_rows, currency_col_indices={7})

    # --- Stream response ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="export_sakti.xlsx"'
    return response


def health_check_view(request):
    """
    Simple health check for Docker and load balancers.
    Returns 200 if the app is up and the database is reachable.
    Note: also defined in config/urls.py for the /health/ route — this version
    lives here so it can be imported by urls.py cleanly.
    """
    try:
        from django.db import connection

        connection.ensure_connection()
        db_status = "ok"
    except Exception as exc:
        return JsonResponse(
            {
                "status": "error",
                "db": str(exc),
                "timestamp": timezone.now().isoformat(),
            },
            status=503,
        )
    return JsonResponse(
        {
            "status": "ok",
            "db": db_status,
            "timestamp": timezone.now().isoformat(),
        }
    )


# ---------------------------------------------------------------------------
# User management (requires authentication)
# ---------------------------------------------------------------------------

@api_view(["GET"])
def users_list_view(request):
    """
    GET /api/users/
    Returns all users sorted by role priority then most-recent last_login.
    Requires a valid JWT access token.
    """
    from django.contrib.auth import get_user_model
    from django.db.models import Case, IntegerField, Value, When

    if not request.user or not request.user.is_authenticated:
        from rest_framework.response import Response as DRFResponse
        return DRFResponse({"detail": "Authentication credentials were not provided."}, status=401)

    User = get_user_model()
    # Role priority: admin first, then operator, then viewer
    role_order = Case(
        When(role="admin",    then=Value(0)),
        When(role="operator", then=Value(1)),
        default=Value(2),
        output_field=IntegerField(),
    )
    qs = (
        User.objects
        .select_related("unit_eselon_i", "unit_eselon_ii")
        .annotate(role_priority=role_order)
        .order_by("role_priority", "-last_login")
    )

    data = []
    for u in qs:
        data.append({
            "id":            u.id,
            "username":      u.username,
            "email":         u.email,
            "display_name":  u.get_full_name() or u.username,
            "first_name":    u.first_name,
            "last_name":     u.last_name,
            "role":          u.role,
            "nip":           u.nip,
            "jabatan":       u.jabatan,
            "unit_eselon_i": {
                "id":    u.unit_eselon_i.id,
                "nama":  u.unit_eselon_i.nama,
                "jenis": u.unit_eselon_i.jenis,
            } if u.unit_eselon_i else None,
            "unit_eselon_ii": {
                "id":   u.unit_eselon_ii.id,
                "nama": u.unit_eselon_ii.nama,
            } if u.unit_eselon_ii else None,
            "last_login":  u.last_login.isoformat() if u.last_login else None,
            "date_joined": u.date_joined.isoformat(),
            "is_active":   u.is_active,
        })
    return Response(data)
